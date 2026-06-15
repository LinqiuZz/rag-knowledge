"""统一摄取管道"""

from __future__ import annotations

import hashlib
import shutil
import traceback
import sys
from pathlib import Path
from typing import TYPE_CHECKING

# 设置控制台编码
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from rich.console import Console

from .pdf import extract_pdf
from .webpage import extract_webpage
from .chunking import split_text
from .chunking_enhanced import split_text as split_text_enhanced

if TYPE_CHECKING:
    from ..config import Settings
    from ..store.vector import VectorStore
    from ..store.metadata import MetadataStore
    from ..store.embedding import EmbeddingManager

console = Console()

# 支持的文件格式
SUPPORTED_EXTENSIONS = {
    # 文档
    '.pdf': 'pdf',
    '.docx': 'word',
    '.doc': 'word',
    '.xlsx': 'excel',
    '.xls': 'excel',
    '.md': 'text',
    '.txt': 'text',
    '.text': 'text',
    # 代码
    '.py': 'code',
    '.js': 'code',
    '.ts': 'code',
    '.java': 'code',
    '.cpp': 'code',
    '.c': 'code',
    '.h': 'code',
    '.hpp': 'code',
    '.go': 'code',
    '.rs': 'code',
    '.rb': 'code',
    '.php': 'code',
    '.swift': 'code',
    '.kt': 'code',
    '.scala': 'code',
    '.r': 'code',
    '.m': 'code',
    '.sql': 'code',
    '.sh': 'code',
    '.bash': 'code',
    '.zsh': 'code',
    '.ps1': 'code',
    '.bat': 'code',
    '.cmd': 'code',
    # 配置/数据
    '.json': 'text',
    '.yaml': 'text',
    '.yml': 'text',
    '.toml': 'text',
    '.ini': 'text',
    '.cfg': 'text',
    '.conf': 'text',
    '.xml': 'text',
    '.html': 'text',
    '.htm': 'text',
    '.css': 'text',
    '.csv': 'text',
}


def _make_id(source: str, chunk_idx: int) -> str:
    """生成确定性块 ID。"""
    h = hashlib.md5(source.encode()).hexdigest()[:12]
    return f"{h}_{chunk_idx:04d}"


def _rollback_vector_store(vector_store: VectorStore, source: str, logger=None):
    """回滚向量库中指定来源的数据。"""
    try:
        deleted = vector_store.delete_by_source(source)
        if deleted > 0:
            console.print(f"  [yellow]⚠ 回滚: 已删除 {deleted} 个向量块[/yellow]")
            if logger:
                logger.log(source, "rollback", f"回滚删除 {deleted} 个向量块")
    except Exception as e:
        console.print(f"  [red]✗ 回滚失败: {e}[/red]")


def ingest_pdf(
    file_path: str | Path,
    settings: Settings,
    vector_store: VectorStore,
    meta_store: MetadataStore,
    embedder: EmbeddingManager,
) -> dict:
    """摄取单个 PDF 文件。"""
    path = Path(file_path).resolve()
    source = str(path)

    # 检查大小
    size_mb = path.stat().st_size / (1024 * 1024)
    if size_mb > settings.ingest.max_file_size_mb:
        raise ValueError(
            f"文件过大: {size_mb:.1f}MB > {settings.ingest.max_file_size_mb}MB"
        )

    console.print(f"[cyan]解析 PDF:[/cyan] {path.name}")

    # 1. 提取文本
    try:
        doc = extract_pdf(path)
    except Exception as e:
        meta_store.log(source, "error", f"PDF解析失败: {e}")
        raise ValueError(f"PDF解析失败: {e}") from e

    if not doc["text"]:
        meta_store.log(source, "error", "PDF未提取到任何文本内容")
        raise ValueError("PDF 未提取到任何文本内容")

    # 2. 分块
    chunks = split_text(
        doc["text"], settings.chunking.size, settings.chunking.overlap
    )
    if not chunks:
        meta_store.log(source, "error", "分块结果为空")
        raise ValueError("分块结果为空")

    console.print(f"  分块完成: {len(chunks)} 块, {doc['char_count']} 字符")

    # 3. 生成嵌入（最耗时，失败需回滚）
    console.print(f"  生成嵌入向量...")
    try:
        embeddings = embedder.embed(chunks)
    except Exception as e:
        meta_store.log(source, "error", f"嵌入生成失败: {e}")
        raise ValueError(f"嵌入生成失败: {e}") from e

    # 4. 存入向量库
    ids = [_make_id(source, i) for i in range(len(chunks))]
    metadatas = [
        {"source": source, "doc_type": "pdf", "title": doc["title"], "chunk_idx": i}
        for i in range(len(chunks))
    ]
    try:
        vector_store.add(ids=ids, documents=chunks, embeddings=embeddings, metadatas=metadatas)
    except Exception as e:
        meta_store.log(source, "error", f"向量存储失败: {e}")
        raise ValueError(f"向量存储失败: {e}") from e

    # 5. 复制原始文件到 data/raw
    try:
        raw_dir = settings.raw_dir / "pdf"
        raw_dir.mkdir(parents=True, exist_ok=True)
        raw_path = raw_dir / path.name
        if not raw_path.exists():
            shutil.copy2(path, raw_path)
    except Exception as e:
        console.print(f"  [yellow]⚠ 原件备份失败（不影响主流程）: {e}[/yellow]")

    # 6. 记录元数据
    try:
        meta_store.add_document(
            source=source, doc_type="pdf", title=doc["title"],
            chunk_count=len(chunks), char_count=doc["char_count"],
        )
        meta_store.log(source, "success", f"摄取完成: {len(chunks)} 块")
    except Exception as e:
        console.print(f"  [yellow]⚠ 元数据记录失败（向量已入库）: {e}[/yellow]")
        meta_store.log(source, "warning", f"元数据记录失败: {e}")

    console.print(f"  [green]✓ 完成:[/green] {doc['title']} ({len(chunks)} 块)")
    return {"title": doc["title"], "chunks": len(chunks), "chars": doc["char_count"]}




def ingest_text(
    file_path: str | Path,
    settings: Settings,
    vector_store: VectorStore,
    meta_store: MetadataStore,
    embedder: EmbeddingManager,
) -> dict:
    """摄取文本/代码文件。支持所有在 SUPPORTED_EXTENSIONS 中标记为 text 或 code 的格式。"""
    path = Path(file_path).resolve()
    source = str(path)

    # 检查文件扩展名
    ext = path.suffix.lower()
    file_type = SUPPORTED_EXTENSIONS.get(ext)
    if file_type not in ('text', 'code'):
        raise ValueError(f"不支持的文件格式: {ext}")
    
    # 检查大小
    size_mb = path.stat().st_size / (1024 * 1024)
    if size_mb > settings.ingest.max_file_size_mb:
        raise ValueError(
            f"文件过大: {size_mb:.1f}MB > {settings.ingest.max_file_size_mb}MB"
        )

    # 确定文档类型
    ext = path.suffix.lower()
    doc_type = SUPPORTED_EXTENSIONS.get(ext, "text")
    type_label = "代码文件" if doc_type == "code" else "文本文件"
    console.print(f"[cyan]解析{type_label}:[/cyan] {path.name}")

    # 1. 读取文本内容
    try:
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    except Exception as e:
        meta_store.log(source, "error", f"文件读取失败: {e}")
        raise ValueError(f"文件读取失败: {e}") from e

    if not text.strip():
        meta_store.log(source, "error", "文件内容为空")
        raise ValueError("文件内容为空")

    # 提取标题（第一行或文件名）
    lines = text.strip().split('\n')
    title = lines[0].lstrip('#').strip() if lines else path.stem
    if not title:
        title = path.stem

    # 2. 分块（使用增强的分块策略）
    chunks = split_text_enhanced(
        text, settings.chunking.size, settings.chunking.overlap,
        strategy=settings.chunking.strategy
    )
    if not chunks:
        meta_store.log(source, "error", "分块结果为空")
        raise ValueError("分块结果为空")

    console.print(f"  分块完成: {len(chunks)} 块, {len(text)} 字符")

    # 3. 生成嵌入（最耗时，失败需回滚）
    console.print(f"  生成嵌入向量...")
    try:
        embeddings = embedder.embed(chunks)
    except Exception as e:
        meta_store.log(source, "error", f"嵌入生成失败: {e}")
        raise ValueError(f"嵌入生成失败: {e}") from e

    # 4. 存入向量库
    ids = [_make_id(source, i) for i in range(len(chunks))]
    metadatas = [
        {"source": source, "doc_type": doc_type, "title": title, "chunk_idx": i}
        for i in range(len(chunks))
    ]
    try:
        vector_store.add(ids=ids, documents=chunks, embeddings=embeddings, metadatas=metadatas)
    except Exception as e:
        meta_store.log(source, "error", f"向量存储失败: {e}")
        raise ValueError(f"向量存储失败: {e}") from e
    
    # 5. 记录元数据
    try:
        meta_store.add_document(
            source=source, doc_type=doc_type, title=title,
            chunk_count=len(chunks), char_count=len(text),
        )
        meta_store.log(source, "success", f"摄取完成: {len(chunks)} 块")
    except Exception as e:
        console.print(f"  [yellow]⚠ 元数据记录失败（向量已入库）: {e}[/yellow]")
        meta_store.log(source, "warning", f"元数据记录失败: {e}")
    
    console.print(f"  [green]✓ 完成:[/green] {title} ({len(chunks)} 块)")
    return {"title": title, "chunks": len(chunks), "chars": len(text)}


def ingest_word(
    file_path: str | Path,
    settings: Settings,
    vector_store: VectorStore,
    meta_store: MetadataStore,
    embedder: EmbeddingManager,
) -> dict:
    """摄取 Word 文件 (.docx)。"""
    path = Path(file_path).resolve()
    source = str(path)

    # 检查扩展名
    if path.suffix.lower() not in ['.docx', '.doc']:
        raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .docx, .doc")

    # 检查大小
    size_mb = path.stat().st_size / (1024 * 1024)
    if size_mb > settings.ingest.max_file_size_mb:
        raise ValueError(f"文件过大: {size_mb:.1f}MB > {settings.ingest.max_file_size_mb}MB")

    console.print(f"[cyan]解析 Word:[/cyan] {path.name}")

    # 1. 提取文本
    try:
        from docx import Document
        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)
        title = path.stem
    except Exception as e:
        meta_store.log(source, "error", f"Word解析失败: {e}")
        raise ValueError(f"Word解析失败: {e}") from e

    if not text.strip():
        meta_store.log(source, "error", "Word未提取到任何文本内容")
        raise ValueError("Word 未提取到任何文本内容")

    # 2. 分块（使用增强的分块策略）
    chunks = split_text_enhanced(
        text, settings.chunking.size, settings.chunking.overlap,
        strategy=settings.chunking.strategy
    )
    if not chunks:
        meta_store.log(source, "error", "分块结果为空")
        raise ValueError("分块结果为空")

    console.print(f"  分块完成: {len(chunks)} 块, {len(text)} 字符")

    # 3. 生成嵌入
    console.print(f"  生成嵌入向量...")
    try:
        embeddings = embedder.embed(chunks)
    except Exception as e:
        meta_store.log(source, "error", f"嵌入生成失败: {e}")
        raise ValueError(f"嵌入生成失败: {e}") from e

    # 4. 存入向量库
    ids = [_make_id(source, i) for i in range(len(chunks))]
    metadatas = [
        {"source": source, "doc_type": "word", "title": title, "chunk_idx": i}
        for i in range(len(chunks))
    ]
    try:
        vector_store.add(ids=ids, documents=chunks, embeddings=embeddings, metadatas=metadatas)
    except Exception as e:
        meta_store.log(source, "error", f"向量存储失败: {e}")
        raise ValueError(f"向量存储失败: {e}") from e

    # 5. 复制原始文件到 data/raw
    try:
        raw_dir = settings.raw_dir / "word"
        raw_dir.mkdir(parents=True, exist_ok=True)
        raw_path = raw_dir / path.name
        if not raw_path.exists():
            shutil.copy2(path, raw_path)
    except Exception as e:
        console.print(f"  [yellow]⚠ 原件备份失败（不影响主流程）: {e}[/yellow]")

    # 6. 记录元数据
    try:
        meta_store.add_document(
            source=source, doc_type="word", title=title,
            chunk_count=len(chunks), char_count=len(text),
        )
        meta_store.log(source, "success", f"摄取完成: {len(chunks)} 块")
    except Exception as e:
        console.print(f"  [yellow]⚠ 元数据记录失败（向量已入库）: {e}[/yellow]")
        meta_store.log(source, "warning", f"元数据记录失败: {e}")

    console.print(f"  [green]✓ 完成:[/green] {title} ({len(chunks)} 块)")
    return {"title": title, "chunks": len(chunks), "chars": len(text)}


def ingest_excel(
    file_path: str | Path,
    settings: Settings,
    vector_store: VectorStore,
    meta_store: MetadataStore,
    embedder: EmbeddingManager,
) -> dict:
    """摄取 Excel 文件 (.xlsx, .xls)。"""
    path = Path(file_path).resolve()
    source = str(path)

    # 检查扩展名
    if path.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .xlsx, .xls")

    # 检查大小
    size_mb = path.stat().st_size / (1024 * 1024)
    if size_mb > settings.ingest.max_file_size_mb:
        raise ValueError(f"文件过大: {size_mb:.1f}MB > {settings.ingest.max_file_size_mb}MB")

    console.print(f"[cyan]解析 Excel:[/cyan] {path.name}")

    # 1. 提取文本
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_text = [f"【{sheet_name}】"]
            for row in ws.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else "" for cell in row]
                if any(row_text):
                    sheet_text.append(" | ".join(row_text))
            all_text.append("\n".join(sheet_text))
        wb.close()
        text = "\n\n".join(all_text)
        title = path.stem
    except Exception as e:
        meta_store.log(source, "error", f"Excel解析失败: {e}")
        raise ValueError(f"Excel解析失败: {e}") from e

    if not text.strip():
        meta_store.log(source, "error", "Excel未提取到任何文本内容")
        raise ValueError("Excel 未提取到任何文本内容")

    # 2. 分块（使用增强的分块策略）
    chunks = split_text_enhanced(
        text, settings.chunking.size, settings.chunking.overlap,
        strategy=settings.chunking.strategy
    )
    if not chunks:
        meta_store.log(source, "error", "分块结果为空")
        raise ValueError("分块结果为空")

    console.print(f"  分块完成: {len(chunks)} 块, {len(text)} 字符")

    # 3. 生成嵌入
    console.print(f"  生成嵌入向量...")
    try:
        embeddings = embedder.embed(chunks)
    except Exception as e:
        meta_store.log(source, "error", f"嵌入生成失败: {e}")
        raise ValueError(f"嵌入生成失败: {e}") from e

    # 4. 存入向量库
    ids = [_make_id(source, i) for i in range(len(chunks))]
    metadatas = [
        {"source": source, "doc_type": "excel", "title": title, "chunk_idx": i}
        for i in range(len(chunks))
    ]
    try:
        vector_store.add(ids=ids, documents=chunks, embeddings=embeddings, metadatas=metadatas)
    except Exception as e:
        meta_store.log(source, "error", f"向量存储失败: {e}")
        raise ValueError(f"向量存储失败: {e}") from e

    # 5. 复制原始文件到 data/raw
    try:
        raw_dir = settings.raw_dir / "excel"
        raw_dir.mkdir(parents=True, exist_ok=True)
        raw_path = raw_dir / path.name
        if not raw_path.exists():
            shutil.copy2(path, raw_path)
    except Exception as e:
        console.print(f"  [yellow]⚠ 原件备份失败（不影响主流程）: {e}[/yellow]")

    # 6. 记录元数据
    try:
        meta_store.add_document(
            source=source, doc_type="excel", title=title,
            chunk_count=len(chunks), char_count=len(text),
        )
        meta_store.log(source, "success", f"摄取完成: {len(chunks)} 块")
    except Exception as e:
        console.print(f"  [yellow]⚠ 元数据记录失败（向量已入库）: {e}[/yellow]")
        meta_store.log(source, "warning", f"元数据记录失败: {e}")

    console.print(f"  [green]✓ 完成:[/green] {title} ({len(chunks)} 块)")
    return {"title": title, "chunks": len(chunks), "chars": len(text)}


def ingest_webpage(
    url: str,
    settings: Settings,
    vector_store: VectorStore,
    meta_store: MetadataStore,
    embedder: EmbeddingManager,
) -> dict:
    """摄取单个网页。"""
    source = url
    console.print(f"[cyan]抓取网页:[/cyan] {url}")

    # 1. 提取正文
    try:
        doc = extract_webpage(url)
    except Exception as e:
        meta_store.log(source, "error", f"网页抓取失败: {e}")
        raise ValueError(f"网页抓取失败: {e}") from e

    if not doc["text"]:
        meta_store.log(source, "error", "网页未提取到任何正文内容")
        raise ValueError("网页未提取到任何正文内容")

    # 2. 分块
    chunks = split_text(
        doc["text"], settings.chunking.size, settings.chunking.overlap
    )
    if not chunks:
        meta_store.log(source, "error", "分块结果为空")
        raise ValueError("分块结果为空")

    console.print(f"  分块完成: {len(chunks)} 块, {doc['char_count']} 字符")

    # 3. 生成嵌入
    console.print(f"  生成嵌入向量...")
    try:
        embeddings = embedder.embed(chunks)
    except Exception as e:
        meta_store.log(source, "error", f"嵌入生成失败: {e}")
        raise ValueError(f"嵌入生成失败: {e}") from e

    # 4. 存入向量库
    ids = [_make_id(source, i) for i in range(len(chunks))]
    metadatas = [
        {"source": source, "doc_type": "webpage", "title": doc["title"], "chunk_idx": i}
        for i in range(len(chunks))
    ]
    try:
        vector_store.add(ids=ids, documents=chunks, embeddings=embeddings, metadatas=metadatas)
    except Exception as e:
        meta_store.log(source, "error", f"向量存储失败: {e}")
        raise ValueError(f"向量存储失败: {e}") from e

    # 5. 记录元数据
    try:
        meta_store.add_document(
            source=source, doc_type="webpage", title=doc["title"],
            chunk_count=len(chunks), char_count=doc["char_count"],
        )
        meta_store.log(source, "success", f"摄取完成: {len(chunks)} 块")
    except Exception as e:
        console.print(f"  [yellow]⚠ 元数据记录失败（向量已入库）: {e}[/yellow]")
        meta_store.log(source, "warning", f"元数据记录失败: {e}")

    console.print(f"  [green]✓ 完成:[/green] {doc['title']} ({len(chunks)} 块)")
    return {"title": doc["title"], "chunks": len(chunks), "chars": doc["char_count"]}
